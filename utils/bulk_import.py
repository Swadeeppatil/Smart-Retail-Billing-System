"""Bulk Product Import from CSV/Excel files"""

import csv
import os
from datetime import datetime


class BulkImporter:
    """Import products from CSV/Excel files"""

    REQUIRED_COLS = ['name', 'category', 'cost_price', 'selling_price', 'stock_quantity']
    OPTIONAL_COLS = ['barcode', 'expiry_date', 'gst_percent']

    def __init__(self, storage, barcode_mgr):
        self.storage = storage
        self.barcode_mgr = barcode_mgr

    def generate_template(self, filepath):
        """Generate a blank CSV template"""
        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(self.REQUIRED_COLS + self.OPTIONAL_COLS)
            writer.writerow(["Sample Product", "Groceries", "50.00", "65.00", "100", "", "", "18"])
        return True

    def import_csv(self, filepath):
        """Import products from CSV. Returns (success_count, errors_list)"""
        if not os.path.exists(filepath):
            return 0, ["File not found"]

        errors = []
        success = 0
        with open(filepath, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row_num, row in enumerate(reader, 2):
                try:
                    name = row.get('name', '').strip()
                    if not name:
                        errors.append(f"Row {row_num}: Missing product name")
                        continue

                    cost = float(row.get('cost_price', 0))
                    sell = float(row.get('selling_price', 0))
                    stock = int(float(row.get('stock_quantity', 0)))
                    category = row.get('category', 'General').strip()
                    barcode = row.get('barcode', '').strip()
                    expiry = row.get('expiry_date', '').strip()
                    gst = float(row.get('gst_percent', 18))

                    if sell <= 0:
                        errors.append(f"Row {row_num}: Invalid selling price for '{name}'")
                        continue

                    if not barcode:
                        barcode = self.barcode_mgr.generate_barcode()

                    from models.models import Product
                    product = Product(
                        name=name, category=category,
                        cost_price=cost, selling_price=sell,
                        stock_quantity=stock, barcode=barcode,
                        expiry_date=expiry, gst_percent=gst
                    )
                    self.storage.add_product(product)
                    success += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

        return success, errors

    def import_excel(self, filepath):
        """Import from Excel file"""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            ws = wb.active

            # Convert to list of dicts
            headers = [cell.value.lower().strip() if cell.value else '' for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: (str(v) if v else '') for i, v in enumerate(row) if i < len(headers)}
                rows.append(row_dict)

            errors = []
            success = 0
            for row_num, row in enumerate(rows, 2):
                try:
                    name = row.get('name', '').strip()
                    if not name:
                        continue
                    cost = float(row.get('cost_price', 0) or 0)
                    sell = float(row.get('selling_price', 0) or 0)
                    stock = int(float(row.get('stock_quantity', 0) or 0))
                    category = row.get('category', 'General').strip()
                    barcode = row.get('barcode', '').strip() or self.barcode_mgr.generate_barcode()
                    expiry = row.get('expiry_date', '').strip()
                    gst = float(row.get('gst_percent', 18) or 18)

                    from models.models import Product
                    product = Product(name=name, category=category, cost_price=cost,
                                      selling_price=sell, stock_quantity=stock,
                                      barcode=barcode, expiry_date=expiry, gst_percent=gst)
                    self.storage.add_product(product)
                    success += 1
                except Exception as e:
                    errors.append(f"Row {row_num}: {str(e)}")

            return success, errors
        except Exception as e:
            return 0, [f"Excel error: {str(e)}"]
