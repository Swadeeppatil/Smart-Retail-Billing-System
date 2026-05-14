"""Excel Export with formatted tables and embedded charts"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils import get_column_letter


class ExcelExporter:
    """Export reports to professional Excel workbooks with charts"""

    HEADER_FILL = PatternFill(start_color="007ACC", end_color="007ACC", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
    TITLE_FONT = Font(bold=True, size=14, color="007ACC")
    BORDER = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def __init__(self, storage, billing_engine=None):
        self.storage = storage
        self.billing = billing_engine

    def export_full_report(self, filepath, from_date=None, to_date=None):
        """Export complete report with multiple sheets and charts"""
        wb = Workbook()

        self._create_sales_sheet(wb, from_date, to_date)
        self._create_products_sheet(wb)
        self._create_summary_sheet(wb, from_date, to_date)

        # Remove default empty sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(filepath)
        return True

    def _style_header(self, ws, row, col_count):
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')
            cell.border = self.BORDER

    def _create_sales_sheet(self, wb, from_date, to_date):
        ws = wb.active
        ws.title = "Sales Report"

        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = "Sales Report"
        ws['A1'].font = self.TITLE_FONT
        ws['A2'] = f"Period: {from_date or 'All'} to {to_date or 'All'}"

        # Headers
        headers = ["Invoice ID", "Date", "Customer", "Items", "Subtotal", "GST", "Discount", "Total"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=4, column=i, value=h)
        self._style_header(ws, 4, len(headers))

        invoices = self.storage.get_all_invoices()
        if from_date:
            invoices = [i for i in invoices if i.created_at[:10] >= from_date]
        if to_date:
            invoices = [i for i in invoices if i.created_at[:10] <= to_date]

        for r, inv in enumerate(invoices, 5):
            ws.cell(row=r, column=1, value=inv.invoice_id)
            ws.cell(row=r, column=2, value=inv.created_at[:10])
            ws.cell(row=r, column=3, value=inv.customer_name)
            ws.cell(row=r, column=4, value=len(inv.items))
            ws.cell(row=r, column=5, value=round(inv.subtotal, 2))
            ws.cell(row=r, column=6, value=round(inv.gst_amount, 2))
            ws.cell(row=r, column=7, value=round(inv.discount_amount, 2))
            ws.cell(row=r, column=8, value=round(inv.grand_total, 2))
            for c in range(1, 9):
                ws.cell(row=r, column=c).border = self.BORDER

        # Auto-width
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Add bar chart
        if invoices:
            chart = BarChart()
            chart.title = "Invoice Totals"
            chart.y_axis.title = "Amount (₹)"
            data = Reference(ws, min_col=8, min_row=4, max_row=4 + len(invoices))
            cats = Reference(ws, min_col=2, min_row=5, max_row=4 + len(invoices))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.width = 20
            ws.add_chart(chart, f"A{6 + len(invoices)}")

    def _create_products_sheet(self, wb):
        ws = wb.create_sheet("Products")
        ws.merge_cells('A1:G1')
        ws['A1'] = "Product Inventory"
        ws['A1'].font = self.TITLE_FONT

        headers = ["Name", "Category", "Cost Price", "Selling Price", "Stock", "Barcode", "Expiry"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=3, column=i, value=h)
        self._style_header(ws, 3, len(headers))

        products = self.storage.get_all_products()
        for r, p in enumerate(products, 4):
            ws.cell(row=r, column=1, value=p.name)
            ws.cell(row=r, column=2, value=p.category)
            ws.cell(row=r, column=3, value=p.cost_price)
            ws.cell(row=r, column=4, value=p.selling_price)
            ws.cell(row=r, column=5, value=p.stock_quantity)
            ws.cell(row=r, column=6, value=p.barcode)
            ws.cell(row=r, column=7, value=p.expiry_date or "-")
            for c in range(1, 8):
                ws.cell(row=r, column=c).border = self.BORDER

        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 15

        # Pie chart for categories
        if products:
            cat_counts = {}
            for p in products:
                cat_counts[p.category] = cat_counts.get(p.category, 0) + 1

            pie_start = 4 + len(products) + 2
            ws.cell(row=pie_start, column=1, value="Category").font = Font(bold=True)
            ws.cell(row=pie_start, column=2, value="Count").font = Font(bold=True)
            for i, (cat, cnt) in enumerate(cat_counts.items(), pie_start + 1):
                ws.cell(row=i, column=1, value=cat)
                ws.cell(row=i, column=2, value=cnt)

            pie = PieChart()
            pie.title = "Products by Category"
            data = Reference(ws, min_col=2, min_row=pie_start, max_row=pie_start + len(cat_counts))
            cats = Reference(ws, min_col=1, min_row=pie_start + 1, max_row=pie_start + len(cat_counts))
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(cats)
            ws.add_chart(pie, f"D{pie_start}")

    def _create_summary_sheet(self, wb, from_date, to_date):
        ws = wb.create_sheet("Summary")
        ws.merge_cells('A1:D1')
        ws['A1'] = "Financial Summary"
        ws['A1'].font = self.TITLE_FONT

        invoices = self.storage.get_all_invoices()
        if from_date:
            invoices = [i for i in invoices if i.created_at[:10] >= from_date]
        if to_date:
            invoices = [i for i in invoices if i.created_at[:10] <= to_date]

        total_rev = sum(i.grand_total for i in invoices)
        total_gst = sum(i.gst_amount for i in invoices)
        total_disc = sum(i.discount_amount for i in invoices)

        stats = [
            ("Total Revenue", f"₹{total_rev:.2f}"),
            ("Total GST Collected", f"₹{total_gst:.2f}"),
            ("Total Discounts Given", f"₹{total_disc:.2f}"),
            ("Total Invoices", str(len(invoices))),
            ("Average Invoice Value", f"₹{total_rev / len(invoices):.2f}" if invoices else "₹0.00"),
        ]

        for i, (label, val) in enumerate(stats, 3):
            ws.cell(row=i, column=1, value=label).font = Font(bold=True)
            ws.cell(row=i, column=2, value=val)
            ws.cell(row=i, column=1).border = self.BORDER
            ws.cell(row=i, column=2).border = self.BORDER

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
